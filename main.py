import csv
import re

import numpy as np
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import openpyxl as op


def replace_datetime_format(text):
    return re.sub(r'(\d{4})-(\d{2})-(\d{2})T(\d{2}:\d{2}:\d{2})\+\d{4}', r'\1', text)

def bold_line(sheet, value):
    work_sheet = sheet[value]
    work_sheet.font = Font(bold=True)

def border_line(sheet, value):
    cell = sheet[value]
    thins = Side(border_style="thin")
    cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)

def bold_border_line(sheet, value):
    bold_line(sheet, value)
    border_line(sheet, value)

def first_salary_dict_sum(i, salary_dict_sum_now, salary_dict_kol_now):
    money_and_long_list = []
    if i[-1] not in salary_dict_sum_now:
        money_and_long_list.append(float(i[1]) + float(i[2]))
        money_and_long_list.append(1)
        salary_dict_sum_now[i[-1]] = money_and_long_list
        salary_dict_kol_now[i[-1]] = 2
    else:
        salary_dict_sum_now[i[-1]][0] = float(salary_dict_sum_now[i[-1]][0]) + float(i[1]) + float(i[2])
        salary_dict_sum_now[i[-1]][1] += 1
        salary_dict_kol_now[i[-1]] += 2

    return salary_dict_sum_now, salary_dict_kol_now

profession = input() # например программист

def create_report():
    excel_doc = op.Workbook()
    excel_doc.create_sheet(title = 'Статистика по годам', index = 0)
    excel_doc.create_sheet(title = 'Статистика по городам', index = 1)
    excel_doc.remove(excel_doc['Sheet'])
    sheetnames = excel_doc.sheetnames
    sheet1 = excel_doc[sheetnames[0]]
    sheet2 = excel_doc[sheetnames[1]]

    salary_dict_sum = {}
    salary_dict_kol = {}
    kol_vacancy = {}
    salary_average = {}
    salary_average_kon = {}
    profession_dict = {}
    all_vacancy = 0

    with open('excel_v.csv', 'r', encoding='utf-8') as csv_reader:
        csv_file = csv.reader(csv_reader)
        for all_vacancy, i in enumerate(csv_file):
            kol_salary_average = []
            profession_three_position = []
            for k, j in enumerate(i):
                i[k] = replace_datetime_format(j)

            salary_dict_now, salary_dict_kol = first_salary_dict_sum(i, salary_dict_sum, salary_dict_kol)

            if not salary_average.get(i[-2]):
                kol_salary_average.append(float(i[1]) + float(i[2]))
                kol_salary_average.append(2.0)
                salary_average[i[-2]] = kol_salary_average
            else:
                salary_average[i[-2]][0] = float(salary_average[i[-2]][0]) + float(i[1]) + float(i[2])
                salary_average[i[-2]][1] += 2.0

            if not kol_vacancy.get(i[-2]):
                kol_vacancy[i[-2]] = 1
            else:
                kol_vacancy[i[-2]] += 1



            if not profession_dict.get(i[-1]) and profession.lower() in i[0].lower():
                profession_three_position.append(float(i[1]) + float(i[2]))
                profession_three_position.append(1)
                profession_dict[i[-1]] = profession_three_position

            elif profession_dict.get(i[-1]) and profession.lower() in i[0].lower():
                profession_dict[i[-1]][0] = profession_dict[i[-1]][0] + float(i[1]) + float(i[2])
                profession_dict[i[-1]][1] += 1


        kon_vacancy_dict = {}

        for city in kol_vacancy:
            if kol_vacancy[city] / all_vacancy > 0.01:
                salary_average_kon[city] = salary_average[city]
                kon_vacancy_dict[city] = kol_vacancy[city]

        for pan, do in enumerate(salary_average_kon):
            salary_average_kon[do] = round(salary_average_kon[do][0] / salary_average_kon[do][1])

        kon_vacancy_dict = sorted(kon_vacancy_dict.items(), key=lambda item: (-item[1], item[0]))
        kon_vacancy_dict = kon_vacancy_dict[:10]
        salary_average_kon = sorted(salary_average_kon.items(), key=lambda x: x[1], reverse=True)
        salary_average_kon = dict(salary_average_kon[:10])

        for i, k in enumerate(salary_dict_sum):
            salary_dict_sum[k][0] = round(salary_dict_sum[k][0] / salary_dict_kol[k])
        salary_dict_sum = sorted(salary_dict_sum.items())

        sheet1.column_dimensions['A'].width = 8
        sheet1['A1'] = 'Год'
        bold_border_line(sheet1, 'A1')
        sheet1.column_dimensions['B'].width = 18
        bold_border_line(sheet1, 'B1')
        sheet1['B1'] = 'Средняя зарплата'
        sheet1.column_dimensions['C'].width = 25
        bold_border_line(sheet1, 'C1')
        sheet1['C1'] = 'Количество вакансий'

        for k, p in enumerate(salary_dict_sum):
            sheet1[f'A{k + 2}'] = salary_dict_sum[k][0]
            border_line(sheet1, f'A{k + 2}')
            sheet1[f'B{k + 2}'] = salary_dict_sum[k][1][0]
            border_line(sheet1, f'B{k + 2}')
            sheet1[f'C{k + 2}'] = salary_dict_sum[k][1][1]
            border_line(sheet1, f'C{k + 2}')

        bold_border_line(sheet2, 'A1')
        sheet2['A1'] = 'Город'
        sheet2.column_dimensions['A'].width = 30
        bold_border_line(sheet2, 'B1')
        sheet2['B1'] = 'Уровень зарплат'
        sheet2.column_dimensions['B'].width = 20

        bold_border_line(sheet2, 'D1')
        sheet2['D1'] = 'Город'
        sheet2.column_dimensions['D'].width = 30
        bold_border_line(sheet2, 'E1')
        sheet2['E1'] = 'Доля вакансий, %'
        sheet2.column_dimensions['E'].width = 20
        dolya_four_graphic = 0
        kon_vacancy_dict_1 = {}
        for gor, zna in enumerate(salary_average_kon):
            sheet2[f'A{gor + 2}'] = zna
            border_line(sheet2, f'A{gor + 2}')
            sheet2[f'B{gor + 2}'] = salary_average_kon[zna]
            border_line(sheet2, f'B{gor + 2}')
            sheet2[f'D{gor + 2}'] = kon_vacancy_dict[gor][0]
            border_line(sheet2, f'D{gor + 2}')
            dolya_four_graphic += round(kon_vacancy_dict[gor][1] / all_vacancy * 100, 2)
            kon_vacancy_dict_1[kon_vacancy_dict[gor][0]] = round(kon_vacancy_dict[gor][1] / all_vacancy * 100, 2)
            sheet2[f'E{gor + 2}'] = round(kon_vacancy_dict[gor][1] / all_vacancy * 100, 2)
            border_line(sheet2, f'E{gor + 2}')

    excel_doc.save('report.xlsx')
    return salary_dict_sum, profession_dict, salary_average_kon, dolya_four_graphic, kon_vacancy_dict_1


def legend_table(plt, year, money, money_profession):
    index = np.arange(len(year))
    bw = 0.4
    plt.set_title('Уровень зарплат по годам', fontsize=8)
    plt.bar(index, money, bw, label='cредняя з/п')
    plt.bar(index + bw, money_profession, bw, label=f'з/п {profession}')
    plt.set_xticks(index + 0.5 * bw, year, fontsize=8, rotation=90)
    plt.tick_params(axis='y', labelsize=8)
    plt.legend(fontsize=8)
    plt.grid(axis='y', linestyle='-', alpha=0.7)

def kol_vacancy(plt, year, kol_all, kol_prof):
    index = np.arange(len(year))
    bw = 0.4
    plt.set_title('Количество вакансий по годам', fontsize=8)
    plt.bar(index, kol_all, bw, label='Количество вакансий')
    plt.bar(index + bw, kol_prof, bw, label=f'Количество вакансий\n{profession}')
    plt.tick_params(axis='y', labelsize=8)
    plt.set_xticks(index + 0.5 * bw, year, fontsize=8, rotation=90)
    plt.legend(fontsize=8)
    plt.grid(axis='y', linestyle='-', alpha=0.7)

def salary_level_by_city(plt1, salary_average_kon):
    city = list(salary_average_kon.keys())
    money = salary_average_kon.values()

    for i, k in enumerate(city):
        city[i] = k.replace('-', '-\n').replace(' ', '\n')


    plt1.set_title('Уровень зарплат по городам', fontsize=8)
    plt1.barh(city, money)
    plt1.tick_params(axis='x', labelsize=8)
    plt1.tick_params(axis='y', labelsize=6)
    plt1.grid(axis='x', linestyle='-')
    plt1.set_yticks(range(len(city)))
    plt1.invert_yaxis()
    plt1.set_yticklabels(city,  va='center', ha='right')



def share_of_vacancies_by_city(plt, kon_vacancy_dict, share_four_graphic):
    kon_vacancy_dict = dict(kon_vacancy_dict)
    name_chart = []
    name_chart.append('Другие')
    name_chart += list(kon_vacancy_dict.keys())
    percent_chart = []
    percent_chart.append(share_four_graphic)
    percent_chart += kon_vacancy_dict.values()
    plt.pie(percent_chart, labels=name_chart, textprops={'fontsize': 6})
    plt.set_title('Доля вакансий по городам', fontsize=8)


def create_plot():
    average_salary = []
    second_graphic = []
    salary_dict_sum, profession_dict, salary_average_kon, share_four_graphic, kon_vacancy_dict = create_report()
    year = list(dict(salary_dict_sum).keys())
    salary_dict_sum = dict(salary_dict_sum)
    year_profession_list = []
    profession_dict_second = {}
    middle_salary_city = []
    for k, i in enumerate(salary_dict_sum):
        average_salary.append(salary_dict_sum[i][0])
        second_graphic.append(salary_dict_sum[i][1])
        if i not in profession_dict:
            profession_dict_second[i] = 0
            profession_dict[i] = [0, 0]
        else:
            profession_dict_second[i] = round(float(profession_dict[i][0]) / (float(profession_dict[i][1] * 2.0)))
    profession_dict_second = dict(sorted(profession_dict_second.items(), key=lambda x: x[0], reverse=False))
    profession_dict = dict(sorted(profession_dict.items(), key=lambda x: x[0], reverse=False))
    for i in profession_dict_second:
        year_profession_list.append(profession_dict_second[i])
    for i in profession_dict:
        middle_salary_city.append(profession_dict[i][1])

    fig, axs = plt.subplots(2, 2)
    (ax1, ax2), (ax3, ax4) = axs
    legend_table(ax1, year, average_salary, year_profession_list)
    kol_vacancy(ax2, year, second_graphic, middle_salary_city)
    salary_level_by_city(ax3, salary_average_kon)
    share_of_vacancies_by_city(ax4, kon_vacancy_dict, share_four_graphic)
    plt.tight_layout()
    plt.show()
    # plt.savefig('C:\\Users\\Notebook\\Desktop\\graph.png', format='png')
    return axs

create_plot()
